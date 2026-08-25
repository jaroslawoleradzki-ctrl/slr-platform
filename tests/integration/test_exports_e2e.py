"""End-to-End integration test suite for Stage 9 Exports & PRISMA Flow (Slice 6).

Exercises the complete real application lifecycle:
1. Full workflow:
   - Project creation
   - Publication import
   - Duplicate group review and canonical merge
   - Title & Abstract screening (with reviewer_a / reviewer_b divergence)
   - Full-Text screening
   - Quality Assessment configuration and execution
   - Data Extraction template configuration and revision submission
2. All 7 export artifacts exercised:
   - BibTeX (.bib)
   - RIS (.ris)
   - XLSX research matrix workbook (.xlsx)
   - CSV extraction dataset (.csv)
   - JSON extraction dataset (.json)
   - PRISMA 2020 SVG diagram (.svg)
   - PRISMA 2020 PDF document (.pdf)
3. Structural guarantees:
   - ZERO superseded records in any export format
   - Reviewer isolation (reviewer_a data vs reviewer_b data)
   - Formula injection neutralization on both data cells and dynamic headers
   - Provenance headers (X-Project-Id, X-Protocol-Version, X-Application-Version, X-Generated-At) on all 7 endpoints
   - In-file D8 provenance comment in BibTeX and RIS matching the response header timestamp
   - Re-import round-trip through production parsers
   - Strictly read-only: persisted SQLite state is 100% identical before and after all export executions
"""

from __future__ import annotations

import io
import sqlite3
import xml.etree.ElementTree as ET
from pathlib import Path
from uuid import uuid4

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.api.main import app
from app.api.routers.exports import get_export_dataset_service
from app.api.routers.extraction import _get_dataset_service
from app.domain.author import Author
from app.domain.extraction import (
    ExtractedValueState,
    ExtractionFieldDefinition,
    ExtractionTemplate,
    ExtractionTemplateVersion,
    FieldDataType,
    ValueOrigin,
    ValueStatus,
)
from app.domain.identifiers import Identifier, IdentifierType
from app.domain.publication import DocumentType, Publication
from app.domain.quality_assessment import (
    QualityAssessmentResponseValue,
    QualityAssessmentTemplate,
    QualityAssessmentTemplateCriterion,
    QualityAssessmentTool,
)
from app.domain.screening import (
    CriterionAssessmentValue,
    ScreeningCriterionStage,
    ScreeningCriterionType,
    ScreeningOutcome,
    ScreeningStage,
)
from app.domain.venue import Venue
from app.providers.import_file.bibtex.mapper import map_bibtex_record
from app.providers.import_file.bibtex.parser import parse_bibtex
from app.providers.import_file.ris.mapper import map_ris_record
from app.providers.import_file.ris.parser import parse_ris
from app.repositories.duplicate_merge_repository import SqliteDuplicateMergeRepository
from app.repositories.duplicate_review_decision_repository import SqliteDuplicateReviewDecisionRepository
from app.repositories.extraction_repository import SqliteExtractionRepository
from app.repositories.extraction_template_repository import SqliteExtractionTemplateRepository
from app.repositories.project_publication_repository import SqliteProjectPublicationRepository
from app.repositories.project_repository import SqliteProjectRepository
from app.repositories.screening_decision_repository import SqliteScreeningDecisionRepository
from app.repositories.sqlite_quality_assessment_repository import (
    SqliteProjectQualityAssessmentConfigurationRepository,
    SqliteQualityAssessmentCatalogRepository,
    SqliteQualityAssessmentRepository,
)
from app.services.export_dataset_service import ExportDatasetService


def _snapshot_database(db_path: Path) -> dict[str, list[tuple]]:
    """Capture all tables and rows in the SQLite database for read-only mutation verification."""
    snapshot: dict[str, list[tuple]] = {}
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()
        tables = [
            row[0]
            for row in cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()
        ]
        for table in tables:
            rows = cursor.execute(f"SELECT * FROM {table} ORDER BY rowid").fetchall()
            snapshot[table] = rows
    return snapshot


@pytest.fixture
def workflow_setup(tmp_path: Path):
    """Execute complete SLR application lifecycle and inject services."""
    db_path = tmp_path / "slr_e2e_workflow.db"

    # 1. Repositories
    project_repo = SqliteProjectRepository(db_path)
    pub_repo = SqliteProjectPublicationRepository(db_path)
    decision_repo = SqliteDuplicateReviewDecisionRepository(db_path)
    merge_repo = SqliteDuplicateMergeRepository(db_path)
    screening_repo = SqliteScreeningDecisionRepository(db_path)
    qa_catalog_repo = SqliteQualityAssessmentCatalogRepository(db_path)
    qa_config_repo = SqliteProjectQualityAssessmentConfigurationRepository(db_path)
    qa_repo = SqliteQualityAssessmentRepository(db_path)
    extraction_template_repo = SqliteExtractionTemplateRepository(db_path)
    extraction_repo = SqliteExtractionRepository(db_path)

    # 2. Project creation via production API / router boundary
    from app.api.dto.project import ProjectCreateRequest
    from app.api.routers.projects import create_project

    project_resp = create_project(
        ProjectCreateRequest(
            title="Comprehensive E2E Lean Energy Study",
            description="End-to-end integration test across all 7 research export formats",
            protocol_version="0.6.0",
        ),
        repo=project_repo,
    )
    project_id = project_resp.project_id

    # 3. Seed publications with deterministic fixed UUIDs
    # Chosen explicitly so that pub2_id < pub1_id, ensuring production merge
    # selects pub2_id as canonical and marks pub1_id as superseded.
    from uuid import UUID

    from app.domain.provenance import ProvenanceEntry

    pub1_id = UUID("bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb")
    pub2_id = UUID("aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa")
    pub3_id = UUID("cccccccc-cccc-4ccc-8ccc-cccccccccccc")

    pub1 = Publication(
        record_id=pub1_id,
        title="Lean energy management in automotive manufacturing (Duplicate Record)",
        authors=[Author(display_name="Kowalski, J.", family_name="Kowalski", given_name="J.")],
        publication_year=2023,
        document_type=DocumentType.JOURNAL_ARTICLE,
        identifiers=[Identifier(type=IdentifierType.DOI, value="10.1016/j.lean.2023.01")],
        provenance=[ProvenanceEntry(source="bibtex", source_record_id=str(pub1_id))],
    )
    pub2 = Publication(
        record_id=pub2_id,
        title="=HYPERLINK(\"http://evil.com\") Lean energy management in automotive manufacturing",
        authors=[Author(display_name="Kowalski, Jan", family_name="Kowalski", given_name="Jan")],
        publication_year=2023,
        document_type=DocumentType.JOURNAL_ARTICLE,
        identifiers=[Identifier(type=IdentifierType.DOI, value="10.1016/j.lean.2023.01")],
        venue=Venue(name="Journal of Cleaner Production"),
        abstract="Investigating lean energy methods in automotive plants.",
        provenance=[ProvenanceEntry(source="bibtex", source_record_id=str(pub2_id))],
    )
    pub3 = Publication(
        record_id=pub3_id,
        title="Zażółć gęślą jaźń: Przegląd efektywności energetycznej",
        authors=[
            Author(display_name="Nowak, Anna", family_name="Nowak", given_name="Anna"),
            Author(display_name="Smith, John", family_name="Smith", given_name="John"),
        ],
        publication_year=2024,
        document_type=DocumentType.JOURNAL_ARTICLE,
        identifiers=[Identifier(type=IdentifierType.DOI, value="10.1000/polish.unicode.2024")],
        venue=Venue(name="Przegląd Elektrotechniczny"),
        abstract="Analiza metod optymalizacji zużycia energii w przemyśle maszynowym.",
        provenance=[ProvenanceEntry(source="bibtex", source_record_id=str(pub3_id))],
    )

    # 3. Seed publications via production ProjectImportService
    from app.api.dto.deduplication import DuplicateDecisionType
    from app.repositories.import_history_repository import SqliteImportHistoryRepository
    from app.repositories.normalization_execution_repository import SqliteNormalizationExecutionRepository
    from app.repositories.transaction_manager import SqliteTransactionManager
    from app.services.project_duplicate_service import ProjectDuplicateService
    from app.services.project_import_service import ProjectImportService

    tx_manager = SqliteTransactionManager(db_path)
    import_service = ProjectImportService(
        publication_repository=pub_repo,
        import_history_repository=SqliteImportHistoryRepository(db_path),
        normalization_repository=SqliteNormalizationExecutionRepository(db_path),
        transaction_manager=tx_manager,
    )
    import_res, _ = import_service.import_bibliographic_publications(
        project_id=project_id,
        filename="lean_manufacturing_studies.bib",
        file_format="bibtex",
        publications=[pub1, pub2, pub3],
    )
    assert import_res.imported_count == 3

    # 4. Execute Duplicate Review & Canonical Merge via production ProjectDuplicateService
    dup_service = ProjectDuplicateService(
        repository=pub_repo,
        decision_repository=decision_repo,
        merge_repository=merge_repo,
        transaction_manager=tx_manager,
    )
    dup_list = dup_service.get_candidate_duplicate_groups(project_id)
    assert dup_list.total_groups_count == 1
    target_group = dup_list.groups[0]

    # Record review decision through production service
    dup_service.record_decision(
        project_id=project_id,
        group_id=target_group.group_id,
        decision=DuplicateDecisionType.APPROVE,
        rationale="Authoritative DOI match and manual verification",
    )

    # Execute canonical merge through production service
    from app.api.dto.deduplication import DuplicateGroupStatus

    merge_res = dup_service.merge_group(project_id=project_id, group_id=target_group.group_id)
    assert merge_res.status == DuplicateGroupStatus.MERGED

    # Crucial: capture authoritative canonical publication ID returned by production merge
    canonical_pub_id = UUID(str(merge_res.canonical_record_id))
    assert canonical_pub_id == pub2_id, "Production merge should select min(UUID) which is pub2_id"
    assert canonical_pub_id != pub1_id

    # 5. Screening Setup (Criteria Reference Data & Decisions via production ScreeningDecisionService)
    from app.domain.screening import (
        ScreeningCriterion,
    )
    from app.repositories.screening_criterion_repository import (
        SqliteScreeningCriterionRepository,
    )
    from app.services.screening_decision_service import (
        CriterionAssessmentInput,
        ScreeningDecisionService,
    )

    screening_criterion_repo = SqliteScreeningCriterionRepository(db_path)
    crit_ta = ScreeningCriterion(
        project_id=project_id,
        name="Topic Relevance",
        description="Topic Relevance criterion",
        criterion_type=ScreeningCriterionType.INCLUSION,
        screening_stage=ScreeningCriterionStage.TITLE_ABSTRACT,
        display_order=1,
        is_active=True,
        is_required=True,
    )
    screening_criterion_repo.create(crit_ta)

    crit_ft = ScreeningCriterion(
        project_id=project_id,
        name="Full Text Quality",
        description="Full Text Quality criterion",
        criterion_type=ScreeningCriterionType.INCLUSION,
        screening_stage=ScreeningCriterionStage.FULL_TEXT,
        display_order=2,
        is_active=True,
        is_required=True,
    )
    screening_criterion_repo.create(crit_ft)

    screening_service = ScreeningDecisionService(
        decision_repository=screening_repo,
        criterion_repository=screening_criterion_repo,
        publication_repository=pub_repo,
    )

    # Reviewer A decisions (INCLUDE canonical_pub_id and pub3 at TA and FT)
    screening_service.record_decision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        stage=ScreeningStage.TITLE_ABSTRACT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_a",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ta.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )
    screening_service.record_decision(
        project_id=project_id,
        publication_id=pub3_id,
        stage=ScreeningStage.TITLE_ABSTRACT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_a",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ta.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )
    screening_service.record_decision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        stage=ScreeningStage.FULL_TEXT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_a",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ft.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )
    screening_service.record_decision(
        project_id=project_id,
        publication_id=pub3_id,
        stage=ScreeningStage.FULL_TEXT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_a",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ft.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )

    # Reviewer B decisions (canonical_pub_id included, pub3 excluded at TA)
    screening_service.record_decision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        stage=ScreeningStage.TITLE_ABSTRACT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_b",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ta.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )
    screening_service.record_decision(
        project_id=project_id,
        publication_id=pub3_id,
        stage=ScreeningStage.TITLE_ABSTRACT,
        outcome=ScreeningOutcome.EXCLUDE,
        reviewer_id="reviewer_b",
        rationale="Out of scope",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ta.criterion_id,
                assessment_value=CriterionAssessmentValue.NOT_MET,
            )
        ],
    )
    screening_service.record_decision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        stage=ScreeningStage.FULL_TEXT,
        outcome=ScreeningOutcome.INCLUDE,
        reviewer_id="reviewer_b",
        assessment_inputs=[
            CriterionAssessmentInput(
                criterion_id=crit_ft.criterion_id,
                assessment_value=CriterionAssessmentValue.MET,
            )
        ],
    )

    # 6. Quality Assessment Configuration & Execution via production services
    from app.services.quality_assessment_configuration_service import (
        DefaultQualityAssessmentConfigurationService,
    )
    from app.services.quality_assessment_execution_service import (
        CriterionResponseInput,
        DefaultQualityAssessmentExecutionService,
    )

    tool_id = "casp_tool"
    qa_catalog_repo.create_tool(
        QualityAssessmentTool(
            tool_id=tool_id,
            name="CASP Quality Assessment",
            description="CASP tool for SLR quality appraisal",
            is_active=True,
        )
    )
    t_version_id = uuid4()
    crit_qa_id = uuid4()
    qa_catalog_repo.create_template_version(
        QualityAssessmentTemplate(
            template_id=t_version_id,
            tool_id=tool_id,
            template_key="casp_v1",
            version=1,
            name="CASP v1",
            description="Version 1",
            is_active=True,
            criteria=[
                QualityAssessmentTemplateCriterion(
                    criterion_id=crit_qa_id,
                    template_id=t_version_id,
                    display_order=1,
                    question="=HYPERLINK(\"http://evil.com\",\"Did study address a clearly focused issue?\")",
                    guidance="Clarity of research aims",
                    is_required=True,
                )
            ],
        )
    )

    qa_config_service = DefaultQualityAssessmentConfigurationService(
        catalog_repo=qa_catalog_repo,
        config_repo=qa_config_repo,
        project_repo=project_repo,
    )
    qa_config_service.configure_project(
        project_id=project_id,
        tool_id=tool_id,
        template_id=t_version_id,
    )

    qa_exec_service = DefaultQualityAssessmentExecutionService(
        project_repo=project_repo,
        publication_repo=pub_repo,
        screening_decision_repo=screening_repo,
        catalog_repo=qa_catalog_repo,
        config_repo=qa_config_repo,
        quality_assessment_repo=qa_repo,
    )

    # Reviewer A QA assessments
    qa_exec_service.save_assessment(
        project_id=project_id,
        publication_id=canonical_pub_id,
        reviewer_id="reviewer_a",
        response_inputs=[
            CriterionResponseInput(
                criterion_id=crit_qa_id,
                response_value=QualityAssessmentResponseValue.YES,
            )
        ],
    )
    qa_exec_service.save_assessment(
        project_id=project_id,
        publication_id=pub3_id,
        reviewer_id="reviewer_a",
        response_inputs=[
            CriterionResponseInput(
                criterion_id=crit_qa_id,
                response_value=QualityAssessmentResponseValue.YES,
            )
        ],
    )

    # Reviewer B QA assessment (canonical_pub_id only)
    qa_exec_service.save_assessment(
        project_id=project_id,
        publication_id=canonical_pub_id,
        reviewer_id="reviewer_b",
        response_inputs=[
            CriterionResponseInput(
                criterion_id=crit_qa_id,
                response_value=QualityAssessmentResponseValue.NO,
                justification="Study methodology lacked focus.",
            )
        ],
    )

    # 7. Data Extraction Template Configuration & Execution via production services
    from app.services.extraction_configuration_service import (
        ExtractionConfigurationService,
    )
    from app.services.extraction_eligibility_service import (
        ExtractionEligibilityService,
        RepositoryQualityAssessmentCompletionReader,
    )
    from app.services.extraction_execution_service import (
        ExtractionExecutionService,
    )

    ext_template_id = "lean_extraction_tmpl"
    ext_version = "1.0.0"
    extraction_template_repo.register_template(
        ExtractionTemplate(template_id=ext_template_id, name="Lean Extraction Template")
    )
    extraction_template_repo.register_version(
        ExtractionTemplateVersion(
            template_id=ext_template_id,
            version=ext_version,
            name="Lean Energy Extraction Template",
            is_active=True,
            is_published=True,
            publication_fields=[
                ExtractionFieldDefinition(
                    field_key="=HYPERLINK(\"http://evil.com\",\"lean_tool\")",
                    name="Lean Tool Category",
                    data_type=FieldDataType.TEXT,
                    description="Category of lean methodology applied",
                ),
                ExtractionFieldDefinition(
                    field_key="energy_savings_pct",
                    name="Energy Savings Percentage",
                    data_type=FieldDataType.TEXT,
                    description="Reported energy consumption reduction percentage",
                ),
            ],
            repeating_groups=[],
        )
    )

    extraction_config_service = ExtractionConfigurationService(
        extraction_repo=extraction_repo,
        template_repo=extraction_template_repo,
        project_repo=project_repo,
    )
    extraction_config_service.set_configuration(
        project_id=project_id,
        template_id=ext_template_id,
        template_version=ext_version,
    )

    from app.services.multi_reviewer_screening_service import (
        MultiReviewerScreeningService,
    )
    from app.services.screening_input_service import (
        ScreeningInputService,
    )

    screening_input_service = ScreeningInputService(
        publication_repository=pub_repo,
        decision_repository=decision_repo,
        merge_repository=merge_repo,
    )
    multi_reviewer_service = MultiReviewerScreeningService(
        input_service=screening_input_service,
    )

    extraction_eligibility_service = ExtractionEligibilityService(
        config_service=extraction_config_service,
        input_service=screening_input_service,
        multi_reviewer_service=multi_reviewer_service,
        decisions_repo=screening_repo,
        qa_completion_reader=RepositoryQualityAssessmentCompletionReader(
            config_repo=qa_config_repo,
            assessment_repo=qa_repo,
        ),
    )
    extraction_exec_service = ExtractionExecutionService(
        config_service=extraction_config_service,
        eligibility_service=extraction_eligibility_service,
        template_repo=extraction_template_repo,
        extraction_repo=extraction_repo,
    )

    # Reviewer A extraction submissions
    extraction_exec_service.submit_revision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        reviewer_id="reviewer_a",
        publication_values=[
            ExtractedValueState(
                field_key="=HYPERLINK(\"http://evil.com\",\"lean_tool\")",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="5S and VSM",
            ),
            ExtractedValueState(
                field_key="energy_savings_pct",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="14.5%",
            ),
        ],
        mark_complete=True,
    )
    extraction_exec_service.submit_revision(
        project_id=project_id,
        publication_id=pub3_id,
        reviewer_id="reviewer_a",
        publication_values=[
            ExtractedValueState(
                field_key="=HYPERLINK(\"http://evil.com\",\"lean_tool\")",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="Kaizen and TPM",
            ),
            ExtractedValueState(
                field_key="energy_savings_pct",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="8.2%",
            ),
        ],
        mark_complete=True,
    )

    # Reviewer B extraction submission (canonical_pub_id only)
    extraction_exec_service.submit_revision(
        project_id=project_id,
        publication_id=canonical_pub_id,
        reviewer_id="reviewer_b",
        publication_values=[
            ExtractedValueState(
                field_key="=HYPERLINK(\"http://evil.com\",\"lean_tool\")",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="Kanban",
            ),
            ExtractedValueState(
                field_key="energy_savings_pct",
                status=ValueStatus.PRESENT,
                origin=ValueOrigin.REVIEWER_CODED,
                reviewer_note="Reviewer note",
                text_value="5.0%",
            ),
        ],
        mark_complete=True,
    )

    # 8. Wire Services and Dependency Overrides
    from app.services.export_dataset_service import _build_extraction_service_for_database

    extraction_dataset_service = _build_extraction_service_for_database(pub_repo, db_path)
    export_service = ExportDatasetService(
        publication_repository=pub_repo,
        project_repository=project_repo,
        extraction_service=extraction_dataset_service,
        qa_catalog_repository=qa_catalog_repo,
        qa_configuration_repository=qa_config_repo,
        qa_repository=qa_repo,
        extraction_template_repository=extraction_template_repo,
    )

    app.dependency_overrides[get_export_dataset_service] = lambda: export_service
    app.dependency_overrides[_get_dataset_service] = lambda: extraction_dataset_service

    client = TestClient(app)
    try:
        yield client, project_id, db_path, pub_repo, project_repo
    finally:
        app.dependency_overrides.pop(get_export_dataset_service, None)
        app.dependency_overrides.pop(_get_dataset_service, None)


class TestExportsEndToEnd:
    """Comprehensive Stage 9 E2E test suite covering all 7 export formats."""

    def test_full_workflow_all_seven_exports_and_provenance_and_readonly(self, workflow_setup):
        test_client, project_id, db_path, _, _ = workflow_setup

        # Step A: Capture persisted SQLite database state BEFORE export execution
        pre_export_snapshot = _snapshot_database(db_path)

        # Step B: Exercise all 7 export endpoints as Reviewer A
        endpoints_reviewer_a = [
            (f"/api/v1/projects/{project_id}/exports/bibtex", "application/x-bibtex", ".bib"),
            (f"/api/v1/projects/{project_id}/exports/ris", "application/x-research-info-systems", ".ris"),
            (f"/api/v1/projects/{project_id}/exports/xlsx?reviewer_id=reviewer_a", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ".xlsx"),
            (f"/api/v1/projects/{project_id}/extraction/export?format=csv&dataset=publications&reviewer_id=reviewer_a", "text/csv", ".csv"),
            (f"/api/v1/projects/{project_id}/extraction/export?format=json&dataset=publications&reviewer_id=reviewer_a", "application/json", None),
            (f"/api/v1/projects/{project_id}/prisma/flow.svg?reviewer_id=reviewer_a", "image/svg+xml", ".svg"),
            (f"/api/v1/projects/{project_id}/prisma/flow.pdf?reviewer_id=reviewer_a", "application/pdf", ".pdf"),
        ]

        responses: dict[str, any] = {}
        for url, expected_media, ext in endpoints_reviewer_a:
            res = test_client.get(url)
            assert res.status_code == 200, f"Failed GET {url}: {res.text}"
            assert expected_media in res.headers.get("content-type", ""), f"Mismatch content-type for {url}"

            # Provenance headers assertion (§16 contract on all 7 formats)
            assert res.headers.get("X-Project-Id") == project_id
            assert res.headers.get("X-Application-Version") == "0.6.4"
            assert "X-Generated-At" in res.headers
            assert res.headers.get("X-Protocol-Version") == "0.6.0"

            if ext:
                disposition = res.headers.get("content-disposition", "")
                assert f"filename=\"{project_id}_" in disposition
                assert disposition.endswith(f"{ext}\"")

            responses[url] = res

        # Step C: Assert D8 in-file provenance on BibTeX and RIS
        bib_res = responses[f"/api/v1/projects/{project_id}/exports/bibtex"]
        ris_res = responses[f"/api/v1/projects/{project_id}/exports/ris"]

        bib_ts = bib_res.headers["X-Generated-At"]
        ris_ts = ris_res.headers["X-Generated-At"]

        assert f"%% Generated by SLR Platform 0.6.4 for project {project_id} at {bib_ts}" in bib_res.text
        assert f"%% Generated by SLR Platform 0.6.4 for project {project_id} at {ris_ts}" in ris_res.text

        # Re-import round-trip
        parsed_bib = parse_bibtex(bib_res.text)
        assert len(parsed_bib) == 2
        mapped_bib = [map_bibtex_record(e, source="bib_test") for e in parsed_bib]
        assert len(mapped_bib) == 2

        parsed_ris = parse_ris(ris_res.text)
        assert len(parsed_ris) == 2
        mapped_ris = [map_ris_record(r, source="ris_test") for r in parsed_ris]
        assert len(mapped_ris) == 2

        # Step D: Assert ZERO superseded duplicate records in any export
        assert "Duplicate Record" not in bib_res.text
        assert "Duplicate Record" not in ris_res.text

        xlsx_res = responses[f"/api/v1/projects/{project_id}/exports/xlsx?reviewer_id=reviewer_a"]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_res.content))
        pub_sheet = wb["Publications"]
        titles = [row[2] for row in pub_sheet.iter_rows(min_row=2, values_only=True) if row[2]]
        assert len(titles) == 2
        assert not any("Duplicate Record" in t for t in titles)

        # Step E: Assert Formula Injection Neutralization in CSV & XLSX (data cells & dynamic headers)
        csv_res = responses[f"/api/v1/projects/{project_id}/extraction/export?format=csv&dataset=publications&reviewer_id=reviewer_a"]
        csv_lines = csv_res.text.splitlines()
        csv_header = csv_lines[0]
        assert "'=HYPERLINK" in csv_header, f"Dynamic header formula was not neutralized: {csv_header}"

        # XLSX formula protection on cells and headers
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    assert cell.data_type != "f", f"Found formula cell in sheet {sheet_name}: {cell.value}"
                    if cell.value and str(cell.value).startswith(("=HYPERLINK", "+SUM", "-SUM", "@SUM")):
                        assert False, f"Unescaped formula cell found in {sheet_name}: {cell.value}"

        pub_title_cell = next(row[2] for row in pub_sheet.iter_rows(min_row=2, values_only=True) if "evil.com" in str(row[2]))
        assert str(pub_title_cell).startswith("'="), f"Title cell was not prefixed with ': {pub_title_cell}"

        ext_sheet = wb["Data Extraction"]
        ext_header_cells = [str(cell.value) for cell in ext_sheet[1]]
        assert any(h.startswith("'=HYPERLINK") for h in ext_header_cells), f"Extraction header was not prefixed with ': {ext_header_cells}"

        # Step F: Assert Reviewer Isolation (Reviewer A vs Reviewer B)
        # Reviewer A has 2 completed extraction records (pub1, pub3)
        json_res_a = responses[f"/api/v1/projects/{project_id}/extraction/export?format=json&dataset=publications&reviewer_id=reviewer_a"]
        records_a = json_res_a.json()
        assert len(records_a) == 2
        assert {r["canonical_title"] for r in records_a} == {
            "=HYPERLINK(\"http://evil.com\") Lean energy management in automotive manufacturing",
            "Zażółć gęślą jaźń: Przegląd efektywności energetycznej",
        }

        # Reviewer B has only 1 completed extraction record (pub1)
        json_res_b = test_client.get(
            f"/api/v1/projects/{project_id}/extraction/export?format=json&dataset=publications&reviewer_id=reviewer_b"
        )
        assert json_res_b.status_code == 200
        records_b = json_res_b.json()
        assert len(records_b) == 1
        assert records_b[0]["canonical_title"] == "=HYPERLINK(\"http://evil.com\") Lean energy management in automotive manufacturing"
        assert records_b[0]["reviewer_id"] == "reviewer_b"

        # PRISMA SVG and PDF for Reviewer A vs Reviewer B
        svg_res_a = responses[f"/api/v1/projects/{project_id}/prisma/flow.svg?reviewer_id=reviewer_a"]
        svg_res_b = test_client.get(f"/api/v1/projects/{project_id}/prisma/flow.svg?reviewer_id=reviewer_b")
        assert svg_res_a.status_code == 200
        assert svg_res_b.status_code == 200

        pdf_res_a = responses[f"/api/v1/projects/{project_id}/prisma/flow.pdf?reviewer_id=reviewer_a"]
        pdf_res_b = test_client.get(f"/api/v1/projects/{project_id}/prisma/flow.pdf?reviewer_id=reviewer_b")
        assert pdf_res_a.status_code == 200
        assert pdf_res_b.status_code == 200
        assert pdf_res_a.content.startswith(b"%PDF-")
        assert pdf_res_b.content.startswith(b"%PDF-")

        # Step G: Assert STRICTLY READ-ONLY PERSISTENCE GUARANTEE
        post_export_snapshot = _snapshot_database(db_path)
        assert pre_export_snapshot == post_export_snapshot, "Database was mutated by export operations!"

    def test_empty_project_returns_valid_empty_artifacts(self, workflow_setup):
        """Ensure empty project produces clean valid artifacts, never 500."""
        from app.api.dto.project import ProjectCreateRequest
        from app.api.routers.projects import create_project

        test_client, _, _, _, project_repo = workflow_setup
        empty_resp = create_project(
            ProjectCreateRequest(
                title="Empty Project",
                description="Empty project test",
                protocol_version="0.6.0",
            ),
            repo=project_repo,
        )
        empty_id = empty_resp.project_id

        # BibTeX -> 200 with D8 comment
        res = test_client.get(f"/api/v1/projects/{empty_id}/exports/bibtex")
        assert res.status_code == 200
        assert "%% Generated by SLR Platform 0.6.4" in res.text

        # RIS -> 200 with D8 comment
        res = test_client.get(f"/api/v1/projects/{empty_id}/exports/ris")
        assert res.status_code == 200
        assert "%% Generated by SLR Platform 0.6.4" in res.text

        # XLSX -> 200 valid workbook with headers
        res = test_client.get(f"/api/v1/projects/{empty_id}/exports/xlsx")
        assert res.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        assert "Publications" in wb.sheetnames

        # PRISMA SVG & PDF -> 200 valid diagram with zeros
        svg_res = test_client.get(f"/api/v1/projects/{empty_id}/prisma/flow.svg")
        assert svg_res.status_code == 200
        assert ET.fromstring(svg_res.text) is not None

        pdf_res = test_client.get(f"/api/v1/projects/{empty_id}/prisma/flow.pdf")
        assert pdf_res.status_code == 200
        assert pdf_res.content.startswith(b"%PDF-")

    def test_unknown_project_returns_404(self, workflow_setup):
        test_client, _, _, _, _ = workflow_setup
        for path in ["exports/bibtex", "exports/ris", "exports/xlsx", "prisma/flow.svg", "prisma/flow.pdf"]:
            res = test_client.get(f"/api/v1/projects/non_existent_project_123/{path}")
            assert res.status_code == 404, f"Expected 404 for {path}, got {res.status_code}"

    def test_frontend_to_backend_stage_9_contract_alignment(self, workflow_setup):
        """Contract-check that the exact HTTP requests constructed by the frontend API map to live backend routes."""
        test_client, project_id, _, _, _ = workflow_setup
        reviewer_id = "reviewer_a"

        # 1. exportApi.exportBibtex -> GET /api/v1/projects/{projectId}/exports/bibtex
        bib_res = test_client.get(f"/api/v1/projects/{project_id}/exports/bibtex")
        assert bib_res.status_code == 200
        assert "application/x-bibtex" in bib_res.headers.get("content-type", "")
        assert bib_res.headers.get("x-project-id") == project_id

        # 2. exportApi.exportRis -> GET /api/v1/projects/{projectId}/exports/ris
        ris_res = test_client.get(f"/api/v1/projects/{project_id}/exports/ris")
        assert ris_res.status_code == 200
        assert "application/x-research-info-systems" in ris_res.headers.get("content-type", "")
        assert ris_res.headers.get("x-project-id") == project_id

        # 3. exportApi.exportXlsx -> GET /api/v1/projects/{projectId}/exports/xlsx?reviewer_id=...
        xlsx_res = test_client.get(f"/api/v1/projects/{project_id}/exports/xlsx?reviewer_id={reviewer_id}")
        assert xlsx_res.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in xlsx_res.headers.get("content-type", "")
        assert xlsx_res.headers.get("x-project-id") == project_id

        # 4. exportApi.exportPrismaSvg -> GET /api/v1/projects/{projectId}/prisma/flow.svg?reviewer_id=...
        svg_res = test_client.get(f"/api/v1/projects/{project_id}/prisma/flow.svg?reviewer_id={reviewer_id}")
        assert svg_res.status_code == 200
        assert "image/svg+xml" in svg_res.headers.get("content-type", "")
        assert svg_res.headers.get("x-project-id") == project_id

        # 5. exportApi.exportPrismaPdf -> GET /api/v1/projects/{projectId}/prisma/flow.pdf?reviewer_id=...
        pdf_res = test_client.get(f"/api/v1/projects/{project_id}/prisma/flow.pdf?reviewer_id={reviewer_id}")
        assert pdf_res.status_code == 200
        assert "application/pdf" in pdf_res.headers.get("content-type", "")
        assert pdf_res.headers.get("x-project-id") == project_id

        # 6. extractionApi.exportDataset(csv) -> GET /api/v1/projects/{projectId}/extraction/export?format=csv&dataset=publications&reviewer_id=...
        csv_res = test_client.get(
            f"/api/v1/projects/{project_id}/extraction/export?format=csv&dataset=publications&reviewer_id={reviewer_id}"
        )
        assert csv_res.status_code == 200
        assert "text/csv" in csv_res.headers.get("content-type", "")
        assert csv_res.headers.get("x-project-id") == project_id

        # 7. extractionApi.exportDataset(json) -> GET /api/v1/projects/{projectId}/extraction/export?format=json&dataset=publications&reviewer_id=...
        json_res = test_client.get(
            f"/api/v1/projects/{project_id}/extraction/export?format=json&dataset=publications&reviewer_id={reviewer_id}"
        )
        assert json_res.status_code == 200
        assert "application/json" in json_res.headers.get("content-type", "")
        assert json_res.headers.get("x-project-id") == project_id
