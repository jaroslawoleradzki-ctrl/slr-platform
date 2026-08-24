import io
from pathlib import Path

import pypdf
import pytest
from fastapi.testclient import TestClient

from app.api.main import app
from app.api.routers.exports import get_export_dataset_service
from app.domain.project import Project
from app.providers.import_file.bibtex.parser import parse_bibtex
from app.providers.import_file.ris.parser import parse_ris
from app.repositories.project_publication_repository import (
    SqliteProjectPublicationRepository,
)
from app.repositories.project_repository import SqliteProjectRepository
from app.services.export_dataset_service import ExportDatasetService
from tests.fixtures.factories import make_publication

PROJECT_ID = "exports_project"


@pytest.fixture
def environment(tmp_path: Path):
    database = tmp_path / "exports.db"
    project_repo = SqliteProjectRepository(database)
    publications = SqliteProjectPublicationRepository(database)
    project_repo.create(Project(project_id=PROJECT_ID, title="Exports Project"))

    service = ExportDatasetService(publication_repository=publications)
    app.dependency_overrides[get_export_dataset_service] = lambda: service
    yield TestClient(app), publications
    app.dependency_overrides.clear()


class TestBibtexEndpoint:
    def test_returns_attachment_with_bibtex_media_type(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex")

        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/x-bibtex")
        assert response.headers["content-disposition"] == 'attachment; filename="exports_project_publications.bib"'
        assert "@misc{" in response.text or "@article{" in response.text

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/missing/exports/bibtex")
        assert response.status_code == 404

    def test_empty_project_yields_valid_empty_artifact(self, environment) -> None:
        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex")
        assert response.status_code == 200
        assert response.text == ""
        assert response.headers["content-disposition"] == 'attachment; filename="exports_project_publications.bib"'

    def test_superseded_records_are_never_exported(self, environment) -> None:
        client, publications = environment
        canonical = make_publication(1)
        superseded = make_publication(2)
        publications.add_publications(PROJECT_ID, [canonical, superseded])
        publications.mark_superseded(PROJECT_ID, [superseded.record_id], canonical.record_id)

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex")

        assert response.status_code == 200
        entries = parse_bibtex(response.text)
        titles = {entry["fields"].get("title") for entry in entries}
        assert titles == {canonical.title}

    def test_two_calls_are_byte_identical(self, environment) -> None:
        client, publications = environment
        publications.add_publications(
            PROJECT_ID,
            [make_publication(1, doi="10.1000/a"), make_publication(2, doi="10.1000/b")],
        )

        first = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex").content
        second = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex").content

        assert first == second

    def test_export_does_not_mutate_project_state(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])
        total_before = publications.count_by_project(PROJECT_ID)
        active_before = publications.count_active_by_project(PROJECT_ID)

        client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex")

        assert publications.count_by_project(PROJECT_ID) == total_before
        assert publications.count_active_by_project(PROJECT_ID) == active_before


class TestRisEndpoint:
    def test_returns_attachment_with_ris_media_type(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris")

        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/x-research-info-systems")
        assert response.headers["content-disposition"] == 'attachment; filename="exports_project_publications.ris"'
        assert "TY  - " in response.text

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/missing/exports/ris")
        assert response.status_code == 404

    def test_empty_project_yields_valid_empty_artifact(self, environment) -> None:
        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris")
        assert response.status_code == 200
        assert response.text == ""

    def test_superseded_records_are_never_exported(self, environment) -> None:
        client, publications = environment
        canonical = make_publication(1)
        superseded = make_publication(2)
        publications.add_publications(PROJECT_ID, [canonical, superseded])
        publications.mark_superseded(PROJECT_ID, [superseded.record_id], canonical.record_id)

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris")

        records = parse_ris(response.text)
        titles = {record["TI"][0] for record in records}
        assert titles == {canonical.title}

    def test_two_calls_are_byte_identical(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])

        first = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris").content
        second = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris").content

        assert first == second

    def test_unicode_titles_survive_the_http_boundary(self, environment) -> None:
        client, publications = environment
        unicode_title = "Wpływ zarządzania lean na efektywność energetyczną"
        publications.add_publications(PROJECT_ID, [make_publication(1, title=unicode_title)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris")

        records = parse_ris(response.text)
        assert records[0]["TI"][0] == unicode_title


class TestXlsxEndpoint:
    def test_returns_attachment_with_xlsx_media_type(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx")

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert response.headers["content-disposition"] == 'attachment; filename="exports_project_publications.xlsx"'
        assert response.content[:2] == b"PK"  # XLSX is a ZIP container

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/missing/exports/xlsx")
        assert response.status_code == 404

    def test_empty_project_yields_workbook_with_all_plan_sheets_headers_only(self, environment) -> None:
        import io

        from openpyxl import load_workbook

        from app.services.export.xlsx_workbook import ALL_SHEET_NAMES, SHEET_PRISMA_SUMMARY

        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx")

        assert response.status_code == 200
        workbook = load_workbook(io.BytesIO(response.content))
        assert tuple(workbook.sheetnames) == ALL_SHEET_NAMES
        for name in ALL_SHEET_NAMES:
            if name == SHEET_PRISMA_SUMMARY:
                assert workbook[name].max_row == 2  # zero-filled authoritative metrics row
            else:
                assert workbook[name].max_row == 1

    def test_publications_sheet_excludes_superseded_records(self, environment) -> None:
        import io

        from openpyxl import load_workbook

        client, publications = environment
        canonical = make_publication(1)
        superseded = make_publication(2)
        publications.add_publications(PROJECT_ID, [canonical, superseded])
        publications.mark_superseded(PROJECT_ID, [superseded.record_id], canonical.record_id)

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx")

        workbook = load_workbook(io.BytesIO(response.content))
        sheet = workbook["Publications"]
        record_ids = {row[0] for row in sheet.iter_rows(min_row=2, values_only=True)}
        assert str(canonical.record_id) in record_ids
        assert str(superseded.record_id) not in record_ids

    def test_unicode_title_survives_the_http_boundary(self, environment) -> None:
        import io

        from openpyxl import load_workbook

        unicode_title = "Wpływ zarządzania lean na efektywność energetyczną"
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1, title=unicode_title)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx")

        workbook = load_workbook(io.BytesIO(response.content))
        titles = {
            row[2] for row in workbook["Publications"].iter_rows(min_row=2, values_only=True)
        }
        assert unicode_title in titles

    def test_export_does_not_mutate_project_state(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])
        total_before = publications.count_by_project(PROJECT_ID)
        active_before = publications.count_active_by_project(PROJECT_ID)

        client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx")

        assert publications.count_by_project(PROJECT_ID) == total_before
        assert publications.count_active_by_project(PROJECT_ID) == active_before

    def test_export_xlsx_accepts_reviewer_id_query_parameter(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/xlsx", params={"reviewer_id": "custom_reviewer"})
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestRoundTripThroughImporterParsers:
    def test_bibtex_export_reimports_with_counts_and_metadata(self, environment) -> None:
        client, publications = environment
        publications.add_publications(
            PROJECT_ID,
            [
                make_publication(1, doi="10.1000/alpha", year=2024),
                make_publication(2, doi="10.1000/beta", year=2023),
                make_publication(3),
            ],
        )

        body = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/bibtex").text
        parsed = parse_bibtex(body)

        assert len(parsed) == 3
        dois = {entry["fields"].get("doi") for entry in parsed}
        assert {"10.1000/alpha", "10.1000/beta"} <= dois

    def test_ris_export_reimports_with_counts_and_metadata(self, environment) -> None:
        client, publications = environment
        publications.add_publications(
            PROJECT_ID,
            [
                make_publication(1, doi="10.1000/alpha", year=2024),
                make_publication(2, doi="10.1000/beta", year=2023),
            ],
        )

        body = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/ris").text
        parsed = parse_ris(body)

        assert len(parsed) == 2
        assert body.count("ER  - \r\n") == 2
        years = sorted(record["PY"][0] for record in parsed)
        assert years == ["2023", "2024"]


class TestPrismaFlowEndpoint:
    def test_returns_200_and_flow_model_json(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/json")

        data = response.json()
        assert data["project_id"] == PROJECT_ID
        assert data["metadata"]["project_title"] == "Exports Project"
        assert len(data["nodes"]) == 7
        assert len(data["edges"]) == 6
        assert "duplicates_removed" in data["removed"]

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/unknown_id/prisma/flow")
        assert response.status_code == 404

    def test_empty_project_returns_zero_counts_flow_model(self, environment) -> None:
        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow")
        assert response.status_code == 200
        data = response.json()
        assert data["metadata"]["counts_echo"]["total_identified"] == 0


class TestPrismaSvgEndpoint:
    def test_returns_attachment_with_svg_media_type(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("image/svg+xml")
        assert response.headers["content-disposition"] == f'attachment; filename="{PROJECT_ID}_prisma_flow.svg"'
        assert "<svg" in response.text
        assert "</svg>" in response.text

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/unknown_id/prisma/flow.svg")
        assert response.status_code == 404

    def test_empty_project_yields_valid_standalone_svg(self, environment) -> None:
        import xml.etree.ElementTree as ET

        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg")
        assert response.status_code == 200
        root = ET.fromstring(response.text)
        assert root.tag.endswith("svg")

    def test_two_calls_are_byte_identical(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])

        first = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg").content
        second = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg").content
        assert first == second

    def test_export_does_not_mutate_project_state(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])
        total_before = publications.count_by_project(PROJECT_ID)

        client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg")
        assert publications.count_by_project(PROJECT_ID) == total_before

    def test_supports_exports_url_alias(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        direct = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.svg").content
        alias = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/prisma/flow.svg").content
        assert direct == alias


class TestPrismaPdfEndpoint:
    def test_returns_attachment_with_pdf_media_type(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("application/pdf")
        assert response.headers["content-disposition"] == f'attachment; filename="{PROJECT_ID}_prisma_flow.pdf"'
        assert response.content.startswith(b"%PDF-")

    def test_unknown_project_returns_404(self, environment) -> None:
        client, _ = environment
        response = client.get("/api/v1/projects/unknown_id/prisma/flow.pdf")
        assert response.status_code == 404

    def test_empty_project_yields_valid_standalone_pdf(self, environment) -> None:
        client, _ = environment
        response = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf")
        assert response.status_code == 200
        assert response.content.startswith(b"%PDF-")

        reader = pypdf.PdfReader(io.BytesIO(response.content))
        assert len(reader.pages) == 1
        text = reader.pages[0].extract_text()
        assert "PRISMA 2020 Flow Diagram" in text
        assert "Active canonical records (n = 0)" in text

    def test_repeated_calls_are_semantically_identical(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])

        first = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf").content
        second = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf").content

        reader1 = pypdf.PdfReader(io.BytesIO(first))
        reader2 = pypdf.PdfReader(io.BytesIO(second))
        assert len(reader1.pages) == len(reader2.pages) == 1
        assert reader1.pages[0].extract_text() == reader2.pages[0].extract_text()

    def test_export_does_not_mutate_project_state(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1), make_publication(2)])
        total_before = publications.count_by_project(PROJECT_ID)

        client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf")
        assert publications.count_by_project(PROJECT_ID) == total_before

    def test_supports_exports_url_alias(self, environment) -> None:
        client, publications = environment
        publications.add_publications(PROJECT_ID, [make_publication(1)])

        direct = client.get(f"/api/v1/projects/{PROJECT_ID}/prisma/flow.pdf").content
        alias = client.get(f"/api/v1/projects/{PROJECT_ID}/exports/prisma/flow.pdf").content
        assert direct.startswith(b"%PDF-")
        assert alias.startswith(b"%PDF-")
        assert len(direct) == len(alias)
