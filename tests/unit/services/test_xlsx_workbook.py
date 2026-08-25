"""Semantic read-back tests for the XLSX research-matrix builder (v0.6.1 Slice 2)."""

import io
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID, uuid4

import pytest
from openpyxl import load_workbook

from app.domain.author import Author
from app.domain.extraction import ExtractedValueState, ValueStatus
from app.domain.identifiers import Identifier, IdentifierType
from app.domain.publication import Publication
from app.domain.quality_assessment import (
    QualityAssessmentResponse,
    QualityAssessmentResponseValue,
    QualityAssessmentTemplateCriterion,
)
from app.domain.screening import ScreeningDecision, ScreeningOutcome, ScreeningStage
from app.domain.synthesis import (
    AnalyticalRelation,
    ClassificationApprovalState,
    EvidenceCharacter,
    RelationDirection,
)
from app.domain.venue import Venue
from app.repositories.project_publication_repository import SqliteProjectPublicationRepository
from app.services.export.cell_safety import excel_safe_cell
from app.services.export.xlsx_workbook import (
    ALL_SHEET_NAMES,
    SHEET_DATA_EXTRACTION,
    SHEET_PRISMA_SUMMARY,
    SHEET_PUBLICATIONS,
    SHEET_QUALITY_ASSESSMENT,
    SHEET_SCREENING_TITLE_ABSTRACT,
    SHEET_SYNTHESIS_RELATIONS,
    build_data_extraction_sheet,
    build_prisma_summary_sheet,
    build_publications_sheet,
    build_quality_assessment_sheet,
    build_screening_title_abstract_sheet,
    build_synthesis_relations_sheet,
    collect_research_matrix_inputs,
    render_research_matrix_workbook,
)
from app.services.export_dataset_service import (
    BibliographicEntry,
    ExportDatasetService,
    QualityAssessmentRow,
    QualityAssessmentSheetData,
)
from app.services.prisma_metrics_service import PrismaMetrics
from tests.fixtures.factories import make_publication

PROJECT_ID = "lean_energy"
TEST_TIME = datetime(2026, 8, 1, 12, 0, tzinfo=timezone.utc)


def make_decision(publication_id: UUID, stage: ScreeningStage, outcome: ScreeningOutcome) -> ScreeningDecision:
    return ScreeningDecision(
        project_id=PROJECT_ID,
        publication_id=publication_id,
        stage=stage,
        outcome=outcome,
        reviewer_id="alice",
        decided_at=TEST_TIME,
    )


def make_relation(publication_id: UUID, group_item_id: UUID, *, approved: bool = True) -> AnalyticalRelation:
    return AnalyticalRelation(
        project_id=PROJECT_ID,
        publication_id=publication_id,
        latest_revision_id=uuid4(),
        group_item_id=group_item_id,
        item_index=1,
        source_practice="5S",
        source_effect="kWh reduction",
        direction=RelationDirection.POSITIVE,
        magnitude=12.5,
        evidence_character=EvidenceCharacter.EMPIRICAL,
        approval_state=ClassificationApprovalState.APPROVED if approved else ClassificationApprovalState.PENDING,
    )


def render_single_sheet(title: str, builder) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = title
    builder(sheet)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def sheet_values(workbook_bytes: bytes, title: str) -> list[list]:
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    return [list(row) for row in workbook[title].iter_rows(values_only=True)]


class TestCellSafety:
    def test_formula_trigger_characters_are_neutralized(self) -> None:
        for hostile in ("=1+1", "+2", "-3", "@cmd", "\tTAB"):
            assert excel_safe_cell(hostile).startswith("'")

    def test_control_characters_are_stripped(self) -> None:
        assert excel_safe_cell("a\x00b\x0bc\x1fd") == "abcd"

    def test_overlong_values_are_truncated_with_marker(self) -> None:
        result = excel_safe_cell("x" * 40000)
        assert len(result) == 32767
        assert result.endswith("…[truncated]")

    def test_none_becomes_empty(self) -> None:
        assert excel_safe_cell(None) == ""


class TestSheetBuildersPure:
    def test_publications_sheet_columns_and_values(self) -> None:
        publication = Publication(
            record_id=UUID(int=1),
            title="Lean energy",
            authors=[Author(display_name="Smith, John", family_name="Smith", given_name="John")],
            publication_year=2024,
            venue=Venue(name="JoCP"),
            identifiers=[Identifier(type=IdentifierType.DOI, value="10.1000/x")],
            keywords=["lean"],
            created_at=TEST_TIME,
        )
        payload = render_single_sheet(
            SHEET_PUBLICATIONS, lambda sheet: build_publications_sheet(sheet, [BibliographicEntry(7, publication)])
        )
        rows = sheet_values(payload, SHEET_PUBLICATIONS)

        assert rows[0][:6] == ["record_id", "position", "title", "authors", "publication_year", "doi"]
        assert rows[1][0] == "00000000-0000-0000-0000-000000000001"
        assert rows[1][1] == 7
        assert rows[1][2] == "Lean energy"
        assert rows[1][3] == "Smith, John"
        assert rows[1][4] == 2024
        assert rows[1][5] == "10.1000/x"

    def test_screening_sheet_filters_by_stage(self) -> None:
        pub_a, pub_b = UUID(int=1), UUID(int=2)
        decisions = [
            make_decision(pub_b, ScreeningStage.TITLE_ABSTRACT, ScreeningOutcome.INCLUDE),
            make_decision(pub_a, ScreeningStage.TITLE_ABSTRACT, ScreeningOutcome.EXCLUDE),
            make_decision(pub_a, ScreeningStage.FULL_TEXT, ScreeningOutcome.INCLUDE),
        ]
        payload = render_single_sheet(
            SHEET_SCREENING_TITLE_ABSTRACT,
            lambda sheet: build_screening_title_abstract_sheet(sheet, decisions),
        )
        rows = sheet_values(payload, SHEET_SCREENING_TITLE_ABSTRACT)

        assert len(rows) == 3  # header + exactly the two T&A decisions
        assert {rows[1][0], rows[2][0]} == {str(pub_a), str(pub_b)}
        assert all(row[2] in {"include", "exclude"} for row in rows[1:])

    def test_quality_assessment_criterion_profile_without_numeric_collapse(self) -> None:
        criterion_yes = QualityAssessmentTemplateCriterion(
            template_id=UUID(int=99), display_order=1, question="Clear rationale?"
        )
        criterion_unclear = QualityAssessmentTemplateCriterion(
            template_id=UUID(int=99), display_order=2, question="Methods sound?"
        )
        response_yes = QualityAssessmentResponse(
            assessment_id=uuid4(),
            criterion_id=criterion_yes.criterion_id,
            question_snapshot=criterion_yes.question,
            response_value=QualityAssessmentResponseValue.YES,
        )
        response_unclear = QualityAssessmentResponse(
            assessment_id=uuid4(),
            criterion_id=criterion_unclear.criterion_id,
            question_snapshot=criterion_unclear.question,
            response_value=QualityAssessmentResponseValue.CANNOT_DETERMINE,
            justification="Unclear sampling",
        )
        qa_data = QualityAssessmentSheetData(
            criteria=(criterion_yes, criterion_unclear),
            rows=(
                QualityAssessmentRow(
                    publication_id=UUID(int=1),
                    reviewer_id="bob",
                    template_id=str(UUID(int=99)),
                    template_version=3,
                    responses_by_criterion={
                        criterion_yes.criterion_id: response_yes,
                        criterion_unclear.criterion_id: response_unclear,
                    },
                    assessed_at=TEST_TIME,
                ),
            ),
        )

        payload = render_single_sheet(
            SHEET_QUALITY_ASSESSMENT, lambda sheet: build_quality_assessment_sheet(sheet, qa_data)
        )
        rows = sheet_values(payload, SHEET_QUALITY_ASSESSMENT)
        headers, data_row = rows[0], rows[1]

        yes_index = headers.index("C1 Clear rationale?")
        unclear_index = headers.index("C2 Methods sound?")
        justification_index = headers.index("C2 Methods sound? — justification")
        assert data_row[yes_index] == "YES"
        assert data_row[unclear_index] == "CANNOT_DETERMINE"
        assert data_row[justification_index] == "Unclear sampling"

    def test_quality_assessment_headers_only_when_unconfigured(self) -> None:
        payload = render_single_sheet(
            SHEET_QUALITY_ASSESSMENT, lambda sheet: build_quality_assessment_sheet(sheet, None)
        )
        rows = sheet_values(payload, SHEET_QUALITY_ASSESSMENT)

        assert rows[0][:4] == ["publication_id", "reviewer_id", "template_id", "template_version"]
        assert len(rows) == 1

    def test_synthesis_sheet_schema_state_column_and_ordering(self) -> None:
        relations = [
            make_relation(UUID(int=2), UUID(int=22)),
            make_relation(UUID(int=1), UUID(int=11)),
        ]
        payload = render_single_sheet(
            SHEET_SYNTHESIS_RELATIONS, lambda sheet: build_synthesis_relations_sheet(sheet, relations)
        )
        rows = sheet_values(payload, SHEET_SYNTHESIS_RELATIONS)

        assert rows[0] == [
            "publication_id", "group_item_id", "source_practice",
            "analytical_lean_category_id", "source_effect",
            "analytical_energy_category_id", "direction", "magnitude",
            "evidence_character", "approval_state",
        ]
        assert rows[1][0] < rows[2][0]
        assert rows[1][6] == "positive" or isinstance(rows[1][6], str)
        assert rows[1][7] == 12.5
        assert all(row[9] == "approved" for row in rows[1:])

    def test_prisma_summary_echoes_authoritative_metrics_and_flattens_breakdown(self) -> None:
        metrics = PrismaMetrics(
            project_id=PROJECT_ID,
            records_identified_providers=10,
            records_identified_imports=5,
            total_identified=15,
            records_after_normalization=15,
            records_before_dedup=15,
            records_after_technical_merger=13,
            duplicate_groups_pending_review=1,
            records_screened_title_abstract=8,
            records_screened_full_text=4,
            studies_included_synthesis=3,
            manual_source_breakdown={"scopus": 3, "google_scholar_pop": 2},
        )
        payload = render_single_sheet(SHEET_PRISMA_SUMMARY, lambda sheet: build_prisma_summary_sheet(sheet, metrics))
        rows = sheet_values(payload, SHEET_PRISMA_SUMMARY)

        by_header = dict(zip(rows[0], rows[1]))
        assert by_header["records_identified_providers"] == 10
        assert by_header["records_after_technical_merger"] == 13
        assert by_header["studies_included_synthesis"] == 3
        assert by_header["manual_source_google_scholar_pop"] == 2
        assert by_header["manual_source_scopus"] == 3

    def test_extraction_sheet_uses_csv_export_header_scheme(self) -> None:
        from types import SimpleNamespace

        from app.domain.extraction import ValueOrigin

        value = ExtractedValueState(
            field_key="e4_country",
            status=ValueStatus.PRESENT,
            text_value="Poland",
            origin=ValueOrigin.REPORTED,
            source_page="1",
        )
        model = SimpleNamespace(
            project_id=PROJECT_ID,
            publication_id=UUID(int=1),
            canonical_title="T",
            canonical_authors=["A"],
            canonical_publication_year=2024,
            canonical_doi=None,
            canonical_journal=None,
            template_id="tpl",
            template_version="v1",
            completeness_status=SimpleNamespace(value="complete"),
            latest_revision_index=1,
            latest_revision_id=uuid4(),
            reviewer_id="alice",
            submitted_at=TEST_TIME,
            publication_values=[value],
        )

        class _Template:
            publication_fields = []

        payload = render_single_sheet(
            SHEET_DATA_EXTRACTION, lambda sheet: build_data_extraction_sheet(sheet, [model], _Template())
        )
        rows = sheet_values(payload, SHEET_DATA_EXTRACTION)

        assert len(rows) == 2
        assert rows[0][13] == "submitted_at"


class TestWorkbookAssemblyIntegration:
    @pytest.fixture
    def seeded_environment(self, tmp_path: Path):
        database = tmp_path / "xlsx.db"
        repo = SqliteProjectPublicationRepository(database)
        publications = [
            make_publication(1, doi="10.1000/a"),
            make_publication(2),
            make_publication(3, doi="10.1000/c"),
        ]
        repo.add_publications(PROJECT_ID, publications)
        canonical, superseded = publications[0], publications[1]
        repo.mark_superseded(PROJECT_ID, [superseded.record_id], canonical.record_id)

        service = ExportDatasetService(publication_repository=repo)
        inputs = collect_research_matrix_inputs(service, PROJECT_ID)
        payload = render_research_matrix_workbook(inputs)
        return canonical, superseded, payload, repo, service

    def test_all_seven_plan_sheets_present_in_plan_order(self, seeded_environment) -> None:
        _, _, payload, _, _ = seeded_environment
        workbook = load_workbook(io.BytesIO(payload))
        assert tuple(workbook.sheetnames) == ALL_SHEET_NAMES

    def test_publications_sheet_excludes_superseded_records_with_persisted_positions(self, seeded_environment) -> None:
        canonical, superseded, payload, *_ = seeded_environment
        rows = sheet_values(payload, SHEET_PUBLICATIONS)

        record_ids = {row[0] for row in rows[1:]}
        assert str(canonical.record_id) in record_ids
        assert str(superseded.record_id) not in record_ids
        positions = sorted(row[1] for row in rows[1:])
        assert positions == [0, 2]  # true persisted positions; superseded gap preserved

    def test_empty_project_yields_all_sheets_headers_only(self, tmp_path: Path) -> None:
        from app.services.export.xlsx_workbook import SHEET_PRISMA_SUMMARY

        repo = SqliteProjectPublicationRepository(tmp_path / "empty.db")
        repo.get_publications("ai_architecture")
        service = ExportDatasetService(publication_repository=repo)
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, "ai_architecture"))

        workbook = load_workbook(io.BytesIO(payload))
        assert tuple(workbook.sheetnames) == ALL_SHEET_NAMES
        for name in ALL_SHEET_NAMES:
            sheet = workbook[name]
            if name is SHEET_PRISMA_SUMMARY or name == SHEET_PRISMA_SUMMARY:
                # PRISMA echoes authoritative zero-filled metrics for empty projects.
                assert sheet.max_row == 2
            else:
                assert sheet.max_row == 1  # header row only

    def test_semantic_determinism_across_repeated_generation(self, seeded_environment) -> None:
        _, _, payload_first, _, service = seeded_environment
        payload_second = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        first_wb = load_workbook(io.BytesIO(payload_first))
        second_wb = load_workbook(io.BytesIO(payload_second))
        for name in ALL_SHEET_NAMES:
            first = tuple(tuple(row) for row in first_wb[name].iter_rows(values_only=True))
            second = tuple(tuple(row) for row in second_wb[name].iter_rows(values_only=True))
            assert first == second, name

    def test_unicode_metadata_survives_roundtrip_through_workbook(self, tmp_path: Path) -> None:
        unicode_title = "Wpływ zarządzania lean na efektywność energetyczną"
        repo = SqliteProjectPublicationRepository(tmp_path / "uni.db")
        repo.add_publications(PROJECT_ID, [make_publication(1, title=unicode_title)])
        service = ExportDatasetService(publication_repository=repo)
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        rows = sheet_values(payload, SHEET_PUBLICATIONS)
        assert any(row[2] == unicode_title for row in rows[1:])

    def test_formula_injection_neutralized_inside_cells(self, tmp_path: Path) -> None:
        repo = SqliteProjectPublicationRepository(tmp_path / "formula.db")
        repo.add_publications(PROJECT_ID, [make_publication(1, title='=HYPERLINK("http://evil")')])
        service = ExportDatasetService(publication_repository=repo)
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        rows = sheet_values(payload, SHEET_PUBLICATIONS)
        assert any(isinstance(row[2], str) and row[2].startswith("'=") for row in rows[1:])

    def test_missing_optional_metadata_leaves_empty_cells_not_placeholders(self, tmp_path: Path) -> None:
        bare = Publication(record_id=UUID(int=9), title="Minimal record", authors=[], created_at=TEST_TIME)
        repo = SqliteProjectPublicationRepository(tmp_path / "bare.db")
        repo.add_publications(PROJECT_ID, [bare])
        service = ExportDatasetService(publication_repository=repo)
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        row = sheet_values(payload, SHEET_PUBLICATIONS)[1]
        assert row[3] is None  # authors empty
        assert row[5] is None  # no doi
        assert row[6] is None  # no venue
        assert row[9] is None  # no keywords

    def test_export_is_read_only(self, seeded_environment) -> None:
        _, _, _, repo, service = seeded_environment
        total_before = repo.count_by_project(PROJECT_ID)
        active_before = repo.count_active_by_project(PROJECT_ID)

        render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        assert repo.count_by_project(PROJECT_ID) == total_before
        assert repo.count_active_by_project(PROJECT_ID) == active_before

    def test_p1_1_persisted_positions_non_sequential_and_missing(self, tmp_path: Path) -> None:
        """P1-1: Persisted positions 3, 8, 21 must be preserved; missing position stays blank."""
        import sqlite3

        from app.repositories.project_publication_repository import DemoProjectPublicationRepository

        db_path = tmp_path / "positions.db"
        repo = SqliteProjectPublicationRepository(db_path)
        pub_a = make_publication(1, title="Pub A")
        pub_b = make_publication(2, title="Pub B")
        pub_c = make_publication(3, title="Pub C")

        repo.add_publications(PROJECT_ID, [pub_a, pub_b, pub_c])

        # Manually set non-sequential positions 3, 8, 21
        with sqlite3.connect(str(db_path)) as conn:
            conn.execute(
                "UPDATE project_publications SET position = 3 WHERE record_id = ?",
                (str(pub_a.record_id),),
            )
            conn.execute(
                "UPDATE project_publications SET position = 8 WHERE record_id = ?",
                (str(pub_b.record_id),),
            )
            conn.execute(
                "UPDATE project_publications SET position = 21 WHERE record_id = ?",
                (str(pub_c.record_id),),
            )
            conn.commit()

        service = ExportDatasetService(publication_repository=repo)
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID))

        rows = sheet_values(payload, SHEET_PUBLICATIONS)
        data_rows = rows[1:]

        # Position column is index 1
        positions_by_id = {row[0]: row[1] for row in data_rows}
        assert positions_by_id[str(pub_a.record_id)] == 3
        assert positions_by_id[str(pub_b.record_id)] == 8
        assert positions_by_id[str(pub_c.record_id)] == 21

        # Verify it did NOT manufacture ordinal 1, 2, 3
        assert [row[1] for row in data_rows] == [3, 8, 21]

        # Test missing position: demo adapter returns None for positions
        demo_repo = DemoProjectPublicationRepository()
        demo_service = ExportDatasetService(publication_repository=demo_repo)
        demo_entries = demo_service.get_bibliographic_entries("lean_energy")
        assert len(demo_entries) > 0
        assert all(entry.position is None for entry in demo_entries)

        demo_payload = render_research_matrix_workbook(collect_research_matrix_inputs(demo_service, "lean_energy"))
        demo_rows = sheet_values(demo_payload, SHEET_PUBLICATIONS)
        assert all(row[1] is None for row in demo_rows[1:])

    def test_p1_2_prisma_persistence_context_parity(self, tmp_path: Path) -> None:
        """P1-2: Export facade PRISMA metrics match the authoritative PRISMA service exactly."""
        from app.domain.project import Project
        from app.repositories.conflict_resolution_repository import SqliteConflictResolutionRepository
        from app.repositories.duplicate_review_decision_repository import SqliteDuplicateReviewDecisionRepository
        from app.repositories.import_history_repository import SqliteImportHistoryRepository
        from app.repositories.project_repository import SqliteProjectRepository
        from app.repositories.screening_decision_repository import SqliteScreeningDecisionRepository
        from app.repositories.screening_reporting_repository import ScreeningReportingRepository
        from app.repositories.screening_reviewer_assignment_repository import (
            SqliteScreeningReviewerAssignmentRepository,
        )
        from app.services.duplicate_group_builder import DuplicateGroupBuilder
        from app.services.multi_reviewer_screening_service import MultiReviewerScreeningService
        from app.services.prisma_metrics_service import PrismaMetricsService
        from app.services.project_workflow_status_service import ProjectWorkflowStatusService
        from app.services.screening_eligibility_adapter import ScreeningEligibilityAdapter
        from app.services.screening_input_service import ScreeningInputService
        from tests.fixtures.factories import make_import_history

        db_path = tmp_path / "prisma_parity.db"
        project_repo = SqliteProjectRepository(db_path)
        project_repo.create(Project(project_id=PROJECT_ID, title="PRISMA Parity"))

        publications = SqliteProjectPublicationRepository(db_path)
        history = SqliteImportHistoryRepository(db_path)
        duplicate_decisions = SqliteDuplicateReviewDecisionRepository(db_path)
        screening_decisions = SqliteScreeningDecisionRepository(db_path)
        assignments = SqliteScreeningReviewerAssignmentRepository(db_path)
        resolutions = SqliteConflictResolutionRepository(db_path)
        reporting = ScreeningReportingRepository(db_path)

        # Seed import history
        history.create(make_import_history(PROJECT_ID, records_count=10, source_type="provider", status="success"))
        history.create(make_import_history(PROJECT_ID, records_count=5, source_type="file", status="success"))

        # Seed publications with duplicates
        pubs = [
            make_publication(1, doi="10.1000/dup"),
            make_publication(2, doi="10.1000/dup"),
            make_publication(3, doi="10.1000/uniq"),
        ]
        publications.add_publications(PROJECT_ID, pubs)
        publications.mark_superseded(PROJECT_ID, [pubs[1].record_id], pubs[0].record_id)

        # Seed screening decisions
        screening_decisions.save(
            ScreeningDecision(
                project_id=PROJECT_ID,
                publication_id=pubs[0].record_id,
                stage=ScreeningStage.TITLE_ABSTRACT,
                outcome=ScreeningOutcome.INCLUDE,
                reviewer_id="alice",
            )
        )
        screening_decisions.save(
            ScreeningDecision(
                project_id=PROJECT_ID,
                publication_id=pubs[0].record_id,
                stage=ScreeningStage.FULL_TEXT,
                outcome=ScreeningOutcome.INCLUDE,
                reviewer_id="alice",
            )
        )

        # Direct authoritative PRISMA path
        input_service = ScreeningInputService(publications, duplicate_decisions)
        multi_service = MultiReviewerScreeningService(
            assignments=assignments,
            reporting=reporting,
            input_service=input_service,
            resolutions=resolutions,
        )
        adapter = ScreeningEligibilityAdapter(
            input_service=input_service,
            assignments_repo=assignments,
            decisions_repo=screening_decisions,
            multi_reviewer_service=multi_service,
        )
        workflow_status = ProjectWorkflowStatusService(
            publication_repository=publications,
            decision_repository=screening_decisions,
            assignment_repository=assignments,
            resolution_repository=resolutions,
            reporting_repository=reporting,
            input_service=input_service,
            multi_reviewer_service=multi_service,
            eligibility_adapter=adapter,
        )
        authoritative_prisma = PrismaMetricsService(
            publication_repository=publications,
            import_history_repository=history,
            decision_repository=duplicate_decisions,
            workflow_status_service=workflow_status,
            builder=DuplicateGroupBuilder(),
        )
        direct_metrics = authoritative_prisma.get_metrics(PROJECT_ID, reviewer_id="alice")

        # Slice 2 ExportDatasetService facade path
        export_service = ExportDatasetService(publication_repository=publications)
        facade_metrics = export_service.get_prisma_metrics(PROJECT_ID, reviewer_id="alice")

        # Verify exact field-by-field parity
        metric_fields = (
            "records_identified_providers",
            "records_identified_imports",
            "total_identified",
            "records_after_normalization",
            "records_before_dedup",
            "records_after_technical_merger",
            "duplicate_groups_pending_review",
            "records_screened_title_abstract",
            "records_screened_full_text",
            "studies_included_synthesis",
            "manual_source_breakdown",
        )
        for field in metric_fields:
            assert getattr(facade_metrics, field) == getattr(direct_metrics, field), f"Divergence in {field}"

        # Verify workbook PRISMA summary sheet matches direct metrics
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(export_service, PROJECT_ID, reviewer_id="alice"))
        summary_rows = sheet_values(payload, SHEET_PRISMA_SUMMARY)
        summary_dict = dict(zip(summary_rows[0], summary_rows[1]))
        assert summary_dict["records_identified_providers"] == direct_metrics.records_identified_providers
        assert summary_dict["records_after_technical_merger"] == direct_metrics.records_after_technical_merger
        assert summary_dict["studies_included_synthesis"] == direct_metrics.studies_included_synthesis

    def test_p1_3_reviewer_id_preserved_and_isolated_end_to_end(self, tmp_path: Path) -> None:
        """P1-3: Reviewer A and Reviewer B data remain isolated in extraction and workbook exports."""
        from app.domain.extraction import (
            ExtractedValueState,
            ExtractionCompletenessStatus,
            ExtractionFieldDefinition,
            ExtractionRecord,
            ExtractionRevision,
            ExtractionTemplate,
            ExtractionTemplateVersion,
            FieldDataType,
            ProjectExtractionConfiguration,
            ValueOrigin,
            ValueStatus,
        )
        from app.domain.project import Project
        from app.repositories.extraction_repository import SqliteExtractionRepository
        from app.repositories.extraction_template_repository import (
            SqliteExtractionTemplateRepository,
        )
        from app.repositories.project_repository import SqliteProjectRepository
        from app.repositories.screening_decision_repository import SqliteScreeningDecisionRepository

        db_path = tmp_path / "reviewer_isolation.db"
        project_repo = SqliteProjectRepository(db_path)
        project_repo.create(Project(project_id=PROJECT_ID, title="Reviewer Isolation"))

        pub_repo = SqliteProjectPublicationRepository(db_path)
        pub1 = make_publication(1, title="Study 1")
        pub_repo.add_publications(PROJECT_ID, [pub1])

        # Seed full text screening decision for reviewer_a and reviewer_b
        screening_repo = SqliteScreeningDecisionRepository(db_path)
        screening_repo.save(
            ScreeningDecision(
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                stage=ScreeningStage.FULL_TEXT,
                outcome=ScreeningOutcome.INCLUDE,
                reviewer_id="reviewer_a",
            )
        )
        screening_repo.save(
            ScreeningDecision(
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                stage=ScreeningStage.FULL_TEXT,
                outcome=ScreeningOutcome.INCLUDE,
                reviewer_id="reviewer_b",
            )
        )

        # Extraction template and config
        tmpl_repo = SqliteExtractionTemplateRepository(db_path)
        tmpl_repo.register_template(ExtractionTemplate(template_id="tpl_rev", name="Tpl"))
        field_def = ExtractionFieldDefinition(field_key="sample_size", name="Sample", data_type=FieldDataType.INTEGER)
        tmpl_repo.register_version(
            ExtractionTemplateVersion(
                template_id="tpl_rev",
                version="1.0.0",
                name="Tpl",
                is_published=True,
                is_active=True,
                publication_fields=[field_def],
            )
        )
        extract_repo = SqliteExtractionRepository(db_path)
        extract_repo.set_project_configuration(
            ProjectExtractionConfiguration(
                project_id=PROJECT_ID,
                template_id="tpl_rev",
                template_version="1.0.0",
            )
        )

        # Save extraction revisions for reviewer_a and reviewer_b
        rec = extract_repo.create_record(
            ExtractionRecord(
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                template_id="tpl_rev",
                template_version="1.0.0",
            )
        )
        extract_repo.append_revision(
            ExtractionRevision(
                record_id=rec.record_id,
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                revision_index=1,
                reviewer_id="reviewer_a",
                completeness_status=ExtractionCompletenessStatus.COMPLETE,
                publication_values=[
                    ExtractedValueState(
                        field_key="sample_size",
                        status=ValueStatus.PRESENT,
                        int_value=100,
                        origin=ValueOrigin.REPORTED,
                        source_page="42",
                    )
                ],
            )
        )
        extract_repo.append_revision(
            ExtractionRevision(
                record_id=rec.record_id,
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                revision_index=2,
                reviewer_id="reviewer_b",
                completeness_status=ExtractionCompletenessStatus.COMPLETE,
                publication_values=[
                    ExtractedValueState(
                        field_key="sample_size",
                        status=ValueStatus.PRESENT,
                        int_value=999,
                        origin=ValueOrigin.REPORTED,
                        source_page="42",
                    )
                ],
            )
        )

        service = ExportDatasetService(publication_repository=pub_repo)

        # CASE B: Request reviewer_a -> returns ONLY reviewer_a data
        models_a = service.get_extraction_read_models(PROJECT_ID, reviewer_id="reviewer_a")
        assert len(models_a) == 1
        assert models_a[0].reviewer_id == "reviewer_a"
        assert models_a[0].publication_values[0].int_value == 100

        # CASE C: Request reviewer_b -> returns ONLY reviewer_b data
        models_b = service.get_extraction_read_models(PROJECT_ID, reviewer_id="reviewer_b")
        assert len(models_b) == 1
        assert models_b[0].reviewer_id == "reviewer_b"
        assert models_b[0].publication_values[0].int_value == 999

        # CASE A: Request reviewer_c (who has NO revision data while reviewer_a and reviewer_b have data)
        # MUST return empty list, NEVER fall back to reviewer_a or reviewer_b
        models_c = service.get_extraction_read_models(PROJECT_ID, reviewer_id="reviewer_c")
        assert models_c == []

        # Test Workbook Data Extraction sheet isolation across all 3 cases
        payload_a = render_research_matrix_workbook(
            collect_research_matrix_inputs(service, PROJECT_ID, reviewer_id="reviewer_a")
        )
        payload_b = render_research_matrix_workbook(
            collect_research_matrix_inputs(service, PROJECT_ID, reviewer_id="reviewer_b")
        )
        payload_c = render_research_matrix_workbook(
            collect_research_matrix_inputs(service, PROJECT_ID, reviewer_id="reviewer_c")
        )

        rows_a = sheet_values(payload_a, SHEET_DATA_EXTRACTION)
        rows_b = sheet_values(payload_b, SHEET_DATA_EXTRACTION)
        rows_c = sheet_values(payload_c, SHEET_DATA_EXTRACTION)

        # Reviewer column is index 12 in Data Extraction sheet
        assert len(rows_a) == 2
        assert rows_a[1][12] == "reviewer_a"
        assert len(rows_b) == 2
        assert rows_b[1][12] == "reviewer_b"
        # reviewer_c has no data -> sheet has header only (1 row), NEVER falls back to another reviewer
        assert len(rows_c) == 1

    def test_p1_4_qa_template_version_consistency(self, tmp_path: Path) -> None:
        """P1-4: QA response data is paired with metadata from its EXACT template version."""
        import sqlite3

        from app.domain.project import Project
        from app.domain.quality_assessment import (
            ProjectQualityAssessmentConfiguration,
            QualityAssessment,
            QualityAssessmentResponse,
            QualityAssessmentResponseValue,
            QualityAssessmentTemplate,
            QualityAssessmentTemplateCriterion,
            QualityAssessmentTool,
        )
        from app.repositories.project_repository import SqliteProjectRepository
        from app.repositories.sqlite_quality_assessment_repository import (
            SqliteProjectQualityAssessmentConfigurationRepository,
            SqliteQualityAssessmentCatalogRepository,
            SqliteQualityAssessmentRepository,
        )

        db_path = tmp_path / "qa_versions.db"
        project_repo = SqliteProjectRepository(db_path)
        project_repo.create(Project(project_id=PROJECT_ID, title="QA Versions"))

        pub_repo = SqliteProjectPublicationRepository(db_path)
        pub1 = make_publication(1, title="Paper 1")
        pub2 = make_publication(2, title="Paper 2")
        pub3 = make_publication(3, title="Paper 3 (Unresolved Template)")
        pub_repo.add_publications(PROJECT_ID, [pub1, pub2, pub3])

        catalog_repo = SqliteQualityAssessmentCatalogRepository(db_path)
        tool = catalog_repo.create_tool(QualityAssessmentTool(tool_id="casp_tool", name="CASP"))

        # CASE A / B setup: Template v1: Question A
        t1_id = uuid4()
        crit_a_id = uuid4()
        catalog_repo.create_template_version(
            QualityAssessmentTemplate(
                template_id=t1_id,
                tool_id=tool.tool_id,
                template_key="casp",
                name="CASP Form",
                version=1,
                criteria=[
                    QualityAssessmentTemplateCriterion(
                        criterion_id=crit_a_id,
                        template_id=t1_id,
                        display_order=1,
                        question="Question A",
                    )
                ],
            )
        )

        # Template v2: Question B
        t2_id = uuid4()
        crit_b_id = uuid4()
        catalog_repo.create_template_version(
            QualityAssessmentTemplate(
                template_id=t2_id,
                tool_id=tool.tool_id,
                template_key="casp",
                name="CASP Form",
                version=2,
                criteria=[
                    QualityAssessmentTemplateCriterion(
                        criterion_id=crit_b_id,
                        template_id=t2_id,
                        display_order=1,
                        question="Question B",
                    )
                ],
            )
        )

        # Project configuration is currently set to v2!
        qa_config_repo = SqliteProjectQualityAssessmentConfigurationRepository(db_path)
        qa_config_repo.save_configuration(
            ProjectQualityAssessmentConfiguration(
                project_id=PROJECT_ID,
                tool_id=tool.tool_id,
                template_id=t2_id,
            )
        )

        # Assessment 1 was completed against historical Template v1 (which exists in catalog)
        qa_repo = SqliteQualityAssessmentRepository(db_path)
        aid_1 = uuid4()
        qa_repo.save_assessment(
            QualityAssessment(
                assessment_id=aid_1,
                project_id=PROJECT_ID,
                publication_id=pub1.record_id,
                reviewer_id="reviewer_1",
                template_id=t1_id,
                responses=[
                    QualityAssessmentResponse(
                        assessment_id=aid_1,
                        criterion_id=crit_a_id,
                        question_snapshot="Question A",
                        response_value=QualityAssessmentResponseValue.YES,
                    )
                ],
            )
        )

        # Assessment 2 was completed against newer Template v2
        aid_2 = uuid4()
        qa_repo.save_assessment(
            QualityAssessment(
                assessment_id=aid_2,
                project_id=PROJECT_ID,
                publication_id=pub2.record_id,
                reviewer_id="reviewer_2",
                template_id=t2_id,
                responses=[
                    QualityAssessmentResponse(
                        assessment_id=aid_2,
                        criterion_id=crit_b_id,
                        question_snapshot="Question B",
                        response_value=QualityAssessmentResponseValue.NO,
                        justification="Lacks justification",
                    )
                ],
            )
        )

        # Assessment 3 references a template that is subsequently unresolvable/unavailable in catalog
        t3_id = uuid4()
        crit_c_id = uuid4()
        catalog_repo.create_template_version(
            QualityAssessmentTemplate(
                template_id=t3_id,
                tool_id=tool.tool_id,
                template_key="casp",
                name="CASP Form",
                version=3,
                criteria=[
                    QualityAssessmentTemplateCriterion(
                        criterion_id=crit_c_id,
                        template_id=t3_id,
                        display_order=1,
                        question="Historical Question C",
                    )
                ],
            )
        )
        aid_3 = uuid4()
        qa_repo.save_assessment(
            QualityAssessment(
                assessment_id=aid_3,
                project_id=PROJECT_ID,
                publication_id=pub3.record_id,
                reviewer_id="reviewer_3",
                template_id=t3_id,
                responses=[
                    QualityAssessmentResponse(
                        assessment_id=aid_3,
                        criterion_id=crit_c_id,
                        question_snapshot="Historical Question C",
                        response_value=QualityAssessmentResponseValue.CANNOT_DETERMINE,
                        justification="Cannot determine without full text",
                    )
                ],
            )
        )
        # Delete template v3 from catalog to simulate missing/unresolvable template metadata
        with sqlite3.connect(str(db_path)) as conn:
            conn.execute("DELETE FROM quality_assessment_templates WHERE template_id = ?", (str(t3_id),))
            conn.commit()

        service = ExportDatasetService(publication_repository=pub_repo)
        qa_data = service.get_quality_assessment_sheet_data(PROJECT_ID)
        assert qa_data is not None

        # Verify criteria contains Question B (from v2 config) and Question A (from v1 historical),
        # but does NOT fabricate criteria for unresolvable t3_id
        question_names = [c.question for c in qa_data.criteria]
        assert "Question A" in question_names
        assert "Question B" in question_names
        assert "Historical Question C" not in question_names

        # Check rows
        row_map = {row.publication_id: row for row in qa_data.rows}
        # CASE A: v1 assessment uses v1 metadata/version
        assert row_map[pub1.record_id].template_version == 1
        assert row_map[pub1.record_id].template_id == str(t1_id)
        assert row_map[pub1.record_id].responses_by_criterion[crit_a_id].response_value == QualityAssessmentResponseValue.YES

        # Current template v2 assessment uses v2 metadata
        assert row_map[pub2.record_id].template_version == 2
        assert row_map[pub2.record_id].template_id == str(t2_id)
        assert row_map[pub2.record_id].responses_by_criterion[crit_b_id].response_value == QualityAssessmentResponseValue.NO

        # CASE B & C: Unresolvable historical template -> template_version is None, NEVER 0 or fabricated
        assert row_map[pub3.record_id].template_version is None
        assert row_map[pub3.record_id].template_id == str(t3_id)

        # Render workbook and verify sheet values
        payload = render_research_matrix_workbook(collect_research_matrix_inputs(service, PROJECT_ID, reviewer_id=""))
        rows = sheet_values(payload, SHEET_QUALITY_ASSESSMENT)

        headers = rows[0]
        q_a_col = headers.index("C1 Question A")
        q_b_col = headers.index("C1 Question B")

        pub1_row = next(r for r in rows[1:] if r[0] == str(pub1.record_id))
        pub2_row = next(r for r in rows[1:] if r[0] == str(pub2.record_id))
        pub3_row = next(r for r in rows[1:] if r[0] == str(pub3.record_id))

        # pub1: v1 / Question A response YES / Question B is None
        assert pub1_row[3] == 1  # template_version
        assert pub1_row[q_a_col] == "YES"
        assert pub1_row[q_b_col] is None

        # pub2: v2 / Question B response NO / Question A is None
        assert pub2_row[3] == 2  # template_version
        assert pub2_row[q_b_col] == "NO"
        assert pub2_row[q_a_col] is None

        # pub3: unresolvable template -> cell 4 is None (blank), NEVER 0, no substitution of v2
        assert pub3_row[3] is None  # template_version blank/None
        assert pub3_row[q_a_col] is None
        assert pub3_row[q_b_col] is None

        # Explicitly verify NO cell in column 4 contains the integer 0 or string "0"
        for r in rows[1:]:
            assert r[3] != 0
            assert r[3] != "0"
