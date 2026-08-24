"""Deterministic XLSX research-matrix workbook builder (v0.6.1 Slice 2, plan §11).

Renders the staged research matrix from persisted project data supplied by
``ExportDatasetService``. The builder is a pure renderer: it owns no queries,
performs no mutation, and never recomputes workflow semantics (PRISMA numbers
are echoed from ``PrismaMetricsService`` output verbatim, D4 untouched).

Sheets — only those backed by persisted models today (plan §11):

1. Publications             — active canonical records with persisted positions
2. Screening Title Abstract — latest recorded T&A decisions on active records
3. Screening Full Text      — latest recorded FT decisions on active records
4. Quality Assessment       — assessed publications, criterion-level profile
                              (no numeric collapse), template-order columns
5. Data Extraction          — Phase 9.8 publication read models, identical
                              header scheme to the extraction CSV export
6. Synthesis Relations      — approved analytical relations of active records
7. PRISMA Summary           — single-row echo of authoritative funnel metrics

Stages without persisted rows are still emitted as header-only sheets so
partially completed reviews remain readable. Every string cell passes through
``excel_safe_cell`` (formula-injection neutralization, control-char stripping,
32 767-character truncation). Absent data is an empty cell — never a
placeholder like "N/A".
"""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.extraction import PublicationExtractionReadModel
from app.domain.screening import ScreeningDecision, ScreeningStage
from app.domain.synthesis import AnalyticalRelation
from app.services.export.cell_safety import excel_safe_cell
from app.services.export_dataset_service import (
    BibliographicEntry,
    ExportDatasetService,
    QualityAssessmentSheetData,
)
from app.services.extraction_dataset_service import _serialize_csv_field, _value_headers
from app.services.prisma_metrics_service import PrismaMetrics

_XLSX_EPOCH = datetime(1970, 1, 1, tzinfo=timezone.utc)

_HEADER_FONT = Font(bold=True)

SHEET_PUBLICATIONS = "Publications"
SHEET_SCREENING_TITLE_ABSTRACT = "Screening Title Abstract"
SHEET_SCREENING_FULL_TEXT = "Screening Full Text"
SHEET_QUALITY_ASSESSMENT = "Quality Assessment"
SHEET_DATA_EXTRACTION = "Data Extraction"
SHEET_SYNTHESIS_RELATIONS = "Synthesis Relations"
SHEET_PRISMA_SUMMARY = "PRISMA Summary"

ALL_SHEET_NAMES = (
    SHEET_PUBLICATIONS,
    SHEET_SCREENING_TITLE_ABSTRACT,
    SHEET_SCREENING_FULL_TEXT,
    SHEET_QUALITY_ASSESSMENT,
    SHEET_DATA_EXTRACTION,
    SHEET_SYNTHESIS_RELATIONS,
    SHEET_PRISMA_SUMMARY,
)


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for column_index in range(1, len(headers) + 1):
        sheet.cell(row=1, column=column_index).font = _HEADER_FONT


def _write_row(sheet: Worksheet, values: list) -> None:
    sheet.append([excel_safe_cell(value) if isinstance(value, str) else value for value in values])


def _iso(value: datetime | None) -> str | None:
    return value.isoformat() if value is not None else None


# ---------------------------------------------------------------------------
# Sheet builders (pure: inputs → worksheet content)
# ---------------------------------------------------------------------------


def _publication_doi(entry: BibliographicEntry) -> str | None:
    from app.domain.identifiers import IdentifierType

    return next(
        (
            identifier.value
            for identifier in entry.publication.identifiers
            if identifier.type is IdentifierType.DOI
        ),
        None,
    )


def build_publications_sheet(sheet: Worksheet, entries: list[BibliographicEntry]) -> None:
    """Active canonical records only; positions are the persisted column values."""
    _write_header(
        sheet,
        [
            "record_id", "position", "title", "authors", "publication_year", "doi",
            "venue", "document_type", "language", "keywords", "urls", "created_at",
        ],
    )
    for entry in entries:
        publication = entry.publication
        _write_row(
            sheet,
            [
                str(publication.record_id),
                entry.position,
                publication.title,
                "; ".join(author.display_name for author in publication.authors),
                publication.publication_year,
                _publication_doi(entry),
                publication.venue.name if publication.venue else None,
                publication.document_type.value if publication.document_type else None,
                publication.language,
                "; ".join(publication.keywords),
                "; ".join(publication.urls),
                _iso(publication.created_at),
            ],
        )


def _build_screening_sheet(sheet: Worksheet, decisions: list[ScreeningDecision], stage: ScreeningStage) -> None:
    _write_header(sheet, ["publication_id", "reviewer_id", "outcome", "decided_at", "rationale"])
    stage_decisions = sorted(
        (decision for decision in decisions if decision.stage is stage),
        key=lambda decision: str(decision.decision_id),
    )
    for decision in stage_decisions:
        _write_row(
            sheet,
            [
                str(decision.publication_id),
                decision.reviewer_id,
                decision.outcome.value,
                _iso(decision.decided_at),
                decision.rationale or None,
            ],
        )


def build_screening_title_abstract_sheet(sheet: Worksheet, decisions: list[ScreeningDecision]) -> None:
    _build_screening_sheet(sheet, decisions, ScreeningStage.TITLE_ABSTRACT)


def build_screening_full_text_sheet(sheet: Worksheet, decisions: list[ScreeningDecision]) -> None:
    _build_screening_sheet(sheet, decisions, ScreeningStage.FULL_TEXT)


def build_quality_assessment_sheet(sheet: Worksheet, qa_data: QualityAssessmentSheetData | None) -> None:
    criterion_headers: list[str] = []
    if qa_data is not None:
        for criterion in qa_data.criteria:
            label = f"C{criterion.display_order} {criterion.question}"
            criterion_headers.append(label)
            criterion_headers.append(f"{label} — justification")
    _write_header(
        sheet,
        ["publication_id", "reviewer_id", "template_id", "template_version", *criterion_headers, "assessed_at"],
    )
    if qa_data is None:
        return
    for row in qa_data.rows:
        values: list = [
            str(row.publication_id),
            row.reviewer_id,
            row.template_id,
            row.template_version,
        ]
        for criterion in qa_data.criteria:
            response = row.responses_by_criterion.get(criterion.criterion_id)
            if response is None:
                values.extend([None, None])
            else:
                justification = response.justification.strip()
                values.append(response.response_value.value)
                values.append(justification if justification else None)
        values.append(_iso(row.assessed_at))
        _write_row(sheet, values)


def build_data_extraction_sheet(
    sheet: Worksheet,
    read_models: list[PublicationExtractionReadModel],
    template,
) -> None:
    """Render extraction read models using the CSV export's header scheme."""
    fields = list(template.publication_fields) if template is not None else []
    headers = [
        "project_id", "publication_id", "canonical_title", "canonical_authors",
        "canonical_publication_year", "canonical_doi", "canonical_journal",
        "template_id", "template_version", "completeness_status",
        "latest_revision_index", "latest_revision_id", "reviewer_id", "submitted_at",
    ] + _value_headers(fields)
    _write_header(sheet, headers)

    for model in read_models:
        row_values: list = [
            model.project_id,
            str(model.publication_id),
            model.canonical_title,
            "; ".join(model.canonical_authors),
            model.canonical_publication_year,
            model.canonical_doi,
            model.canonical_journal,
            model.template_id,
            model.template_version,
            model.completeness_status.value,
            model.latest_revision_index,
            str(model.latest_revision_id),
            model.reviewer_id,
            _iso(model.submitted_at),
        ]
        field_values = {value.field_key: value for value in model.publication_values}
        for field in fields:
            # Identical serialization contract to the extraction CSV export;
            # empty strings become truly empty cells.
            row_values.extend(cell if cell != "" else None for cell in _serialize_csv_field(
                field_values.get(field.field_key), field.data_type
            ))
        _write_row(sheet, row_values)


def build_synthesis_relations_sheet(sheet: Worksheet, relations: list[AnalyticalRelation]) -> None:
    _write_header(
        sheet,
        [
            "publication_id", "group_item_id", "source_practice",
            "analytical_lean_category_id", "source_effect",
            "analytical_energy_category_id", "direction", "magnitude",
            "evidence_character", "approval_state",
        ],
    )
    ordered = sorted(relations, key=lambda relation: (str(relation.publication_id), str(relation.group_item_id)))
    for relation in ordered:
        _write_row(
            sheet,
            [
                str(relation.publication_id),
                str(relation.group_item_id),
                relation.source_practice,
                relation.analytical_lean_category_id,
                relation.source_effect,
                relation.analytical_energy_category_id,
                relation.direction.value,
                relation.magnitude,
                relation.evidence_character.value,
                relation.approval_state.value,
            ],
        )


PRISMA_SUMMARY_METRICS = (
    "records_identified_providers",
    "records_identified_imports",
    "total_identified",
    "records_after_normalization",
    "records_before_dedup",
    "records_after_technical_merger",
    "duplicate_groups_pending_review",
    "records_screened_title_abstract",
    "records_screened_full_text",
    "studies_included_synthesis",
)


def build_prisma_summary_sheet(sheet: Worksheet, metrics: PrismaMetrics) -> None:
    breakdown_keys = sorted(metrics.manual_source_breakdown)
    _write_header(
        sheet,
        ["project_id", *PRISMA_SUMMARY_METRICS, *(f"manual_source_{key}" for key in breakdown_keys)],
    )
    values: list = [metrics.project_id]
    values.extend(getattr(metrics, name) for name in PRISMA_SUMMARY_METRICS)
    values.extend(metrics.manual_source_breakdown[key] for key in breakdown_keys)
    _write_row(sheet, values)


# ---------------------------------------------------------------------------
# Workbook assembly
# ---------------------------------------------------------------------------


def collect_research_matrix_inputs(
    service: ExportDatasetService,
    project_id: str,
    *,
    reviewer_id: str = "default_reviewer",
):
    """Gather all persisted datasets for the workbook via the facade.

    The initial bibliographic read raises ``ProjectNotFoundError`` for unknown
    projects (mapped to HTTP 404 by the router); every later read is safe for
    existing-but-empty projects.
    """
    from app.services.extraction_dataset_service import ExtractionConfigurationNotFoundError

    entries = service.get_bibliographic_entries(project_id)
    screening_decisions = service.get_screening_decisions(project_id, reviewer_id)
    qa_data = service.get_quality_assessment_sheet_data(project_id, reviewer_id=reviewer_id)
    extraction_template = service.get_extraction_template(project_id)
    try:
        extraction_models = service.get_extraction_read_models(project_id, reviewer_id=reviewer_id)
    except ExtractionConfigurationNotFoundError:
        extraction_models = []
    synthesis_relations = service.get_approved_synthesis_relations(project_id)
    prisma_metrics = service.get_prisma_metrics(project_id, reviewer_id)
    return {
        "entries": entries,
        "screening_decisions": screening_decisions,
        "qa_data": qa_data,
        "extraction_template": extraction_template,
        "extraction_models": extraction_models,
        "synthesis_relations": synthesis_relations,
        "prisma_metrics": prisma_metrics,
    }


def render_research_matrix_workbook(inputs: dict) -> bytes:
    """Render the §11 workbook from collected inputs (pure, no repository access)."""
    workbook = Workbook()
    workbook.properties.creator = "SLR Platform"
    workbook.properties.created = _XLSX_EPOCH
    workbook.properties.modified = _XLSX_EPOCH

    publications = workbook.active
    assert publications is not None
    publications.title = SHEET_PUBLICATIONS
    build_publications_sheet(publications, inputs["entries"])

    build_screening_title_abstract_sheet(workbook.create_sheet(SHEET_SCREENING_TITLE_ABSTRACT), inputs["screening_decisions"])
    build_screening_full_text_sheet(workbook.create_sheet(SHEET_SCREENING_FULL_TEXT), inputs["screening_decisions"])
    build_quality_assessment_sheet(workbook.create_sheet(SHEET_QUALITY_ASSESSMENT), inputs["qa_data"])
    build_data_extraction_sheet(
        workbook.create_sheet(SHEET_DATA_EXTRACTION),
        inputs["extraction_models"],
        inputs["extraction_template"],
    )
    build_synthesis_relations_sheet(workbook.create_sheet(SHEET_SYNTHESIS_RELATIONS), inputs["synthesis_relations"])
    build_prisma_summary_sheet(workbook.create_sheet(SHEET_PRISMA_SUMMARY), inputs["prisma_metrics"])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def build_research_matrix_workbook(
    service: ExportDatasetService,
    project_id: str,
    *,
    reviewer_id: str = "default_reviewer",
) -> bytes:
    """Collect persisted datasets and render the research-matrix workbook."""
    return render_research_matrix_workbook(collect_research_matrix_inputs(service, project_id, reviewer_id=reviewer_id))
